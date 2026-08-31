"""归档库多 sheet 覆盖导入驱动。

用法：
  python scripts/bulk_import_archive.py --dry-run     # 仅解析+统计，不写库
  python scripts/bulk_import_archive.py --execute     # 真正覆盖导入

策略：保留 40 个存量项目的身份（project_code/requirement_id/requirement_name 不变），
仅清空并重建其子数据（modules/fps/sub_processes）；额外的真实汇总 sheet 作为新项目追加。
数据源用打过补丁的副本（原文件 styles.xml 触发 openpyxl Fill bug，已剥离 <fills> 不影响数值）。
"""
import argparse
import os
import sys

# 让脚本能 import app 包（脚本在 scripts/ 下，需把仓库根加入 path）
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# 备份目录（dump_dimension_backup 写这里）
BACKUP_DIR = r"C:\Users\13871\WorkBuddy\2026-08-30-22-14-43\backup"
os.makedirs(BACKUP_DIR, exist_ok=True)
os.environ.setdefault("DATA_DIR", BACKUP_DIR)

from openpyxl import load_workbook
from app.services.xlsx_import import bulk_import_workbook, parse_worksheet, group_tree

PATCHED = r"C:\Users\13871\WorkBuddy\2026-08-30-22-14-43\cosmic_patched.xlsx"
DIM = "cosmic_archive"

# 真实 COSMIC 汇总 sheet（排除 7 个跟踪/列表页与示例/空/补录跟踪/汇总页）
# 经结构探测确定：周期性汇总 [7]~[46] 共 40 个 + [48] 二季度线下工作项补录（620 行真实数据）
SUPPLEMENTARY = ["二季度线下工作项补录需求cosmic汇总"]  # 作为新项目追加


def build_mapping():
    from app import db
    existing = db.query(DIM, "SELECT id, requirement_id, requirement_name FROM projects ORDER BY id")
    ex_by_name = { (p["requirement_name"] or "").strip(): p for p in existing }

    wb = load_workbook(PATCHED, data_only=True)
    sheet_names = set(wb.sheetnames)
    mapping = []
    matched, unmatched = [], []
    for p in existing:
        name = (p["requirement_name"] or "").strip()
        if name in sheet_names:
            mapping.append({"sheet": name, "project_id": p["id"]})
            matched.append(name)
        else:
            unmatched.append((p["id"], name))
    # 补充 sheet 作为新项目（REF 顺延）。
    # 注意：若该 sheet 已通过 requirement_name 命中存量项目，则跳过，避免重复建项目。
    max_ref = 0
    for p in existing:
        rid = p["requirement_id"] or ""
        if rid.startswith("REF-"):
            try:
                max_ref = max(max_ref, int(rid[4:]))
            except ValueError:
                pass
    for name in SUPPLEMENTARY:
        if name not in sheet_names:
            continue
        if name in ex_by_name:
            print(f"   SKIP supplementary {name!r}: 已命中存量项目 id={ex_by_name[name]['id']}，不重复建")
            continue
        max_ref += 1
        mapping.append({
            "sheet": name,
            "project_meta": {
                "project_code": "ref_" + name[:12],
                "client_name": "",
                "requirement_id": f"REF-{max_ref}",
                "requirement_name": name,
            },
        })
        matched.append(name + " (NEW)")
    wb.close()
    print(f"[mapping] existing={len(existing)} matched={len(matched)} unmatched={len(unmatched)}")
    for uid, un in unmatched:
        print(f"   UNMATCHED project id={uid} name={un!r}")
    return mapping


def dry_run(mapping):
    wb = load_workbook(PATCHED, data_only=True)
    tot_m = tot_f = tot_s = 0
    print(f"{'sheet':<42}{'mod':>6}{'fp':>7}{'sub':>7}")
    for it in mapping:
        name = it["sheet"]
        parsed = parse_worksheet(wb[name])
        if not parsed["rows"]:
            print(f"{name:<42}  (no data rows)")
            continue
        mods = group_tree(parsed["rows"])
        m = len(mods); f = sum(len(x["fps"]) for x in mods); s = sum(len(y["subs"]) for x in mods for y in x["fps"])
        tot_m += m; tot_f += f; tot_s += s
        print(f"{name:<42}{m:>6}{f:>7}{s:>7}")
    print(f"{'TOTAL':<42}{tot_m:>6}{tot_f:>7}{tot_s:>7}")
    wb.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--execute", action="store_true", help="真正写库（默认 dry-run）")
    args = ap.parse_args()

    mapping = build_mapping()
    if not mapping:
        print("mapping 为空，退出"); sys.exit(1)

    if not args.execute:
        print("=== DRY RUN（不写库）===")
        dry_run(mapping)
        print("\n加 --execute 执行真正覆盖导入。")
        return

    print("=== EXECUTE 覆盖导入 ===")
    rep = bulk_import_workbook(DIM, PATCHED, mapping, backup_tag="pre_bulk_overwrite")
    print("backup:", rep.get("backup"))
    print("created:", rep["created"])
    print("updated:", rep["updated"])
    print("skipped:", rep["skipped"])
    print("errors:", rep["errors"])
    print(f"details: {len(rep['details'])} projects")
    tot_s = sum(d["subs"] for d in rep["details"])
    tot_f = sum(d["fps"] for d in rep["details"])
    tot_m = sum(d["modules"] for d in rep["details"])
    print(f"TOTAL imported -> modules={tot_m} fps={tot_f} subs={tot_s}")


if __name__ == "__main__":
    main()
