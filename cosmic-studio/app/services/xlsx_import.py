"""xlsx 导入：增量 upsert / 全量覆盖，合并单元格传播逻辑移植 import_from_xlsx.py。

列映射（与 Hermes 导入脚本一致）：A=客户需求(不入库,仅覆盖模式解析项目信息)，
B/C/D=一二三级模块，E=功能用户，F=触发事件，G=功能过程，H=描述，I=数据移动，
J=数据组，K=数据属性；数据从第 5 行起，L-P 展示列不导入。

业务主键：模块=(project, level1+level2+level3)；FP=(module, fp_name)；
FP 命中主键时整体重写 functional_user/trigger_event + 子过程全删重插（防半新半旧）。
增量=按主键 upsert 保留库内其他数据（必须指定 project_id）；
覆盖=先备份再清（指定项目或整库）重灌，一个 xlsx = 一个项目。
"""
import io
import re
import zipfile

from openpyxl import load_workbook

from .. import db

REQ_LABEL_RE = re.compile(r"^【(.+?)】(.+)$")


def parse_worksheet(ws) -> dict:
    """→ {requirement_label, rows:[{level1..3,user,event,fp_name,description,move_type,group_name,attributes}]}。
    从已打开的 worksheet 解析（支持指定 sheet / 多 sheet 批量导入）。"""
    prev = {c: "" for c in range(1, 17)}
    rows = []
    req_label = ""
    for r in range(5, ws.max_row + 1):
        vals = {}
        all_none = True
        for c in range(1, 12):
            v = _merged_value(ws, r, c, prev[c])
            prev[c] = v
            if v:
                all_none = False
            vals[c] = v
        if all_none:
            continue
        if not req_label and vals[1]:
            req_label = vals[1]
        rows.append({
            "level1": vals[2], "level2": vals[3], "level3": vals[4],
            "user": vals[5], "event": vals[6], "fp_name": vals[7],
            "description": vals[8], "move_type": (vals[9] or "").strip().upper()[:1],
            "group_name": vals[10], "attributes": vals[11],
        })
    return {"requirement_label": req_label, "rows": rows}


def parse_xlsx_rows(xlsx_bytes: bytes, sheet=None) -> dict:
    """→ {requirement_label, rows:[...]}。sheet=None 时取活动表（兼容旧调用 / 单项目导入）。"""
    wb = load_workbook_patched(xlsx_bytes)
    try:
        ws = wb[sheet] if sheet else wb.active
        return parse_worksheet(ws)
    finally:
        wb.close()


# ---------- 多 sheet 批量覆盖导入（一个工作簿 = 多个项目） ----------

def _repair_xlsx_bytes(xlsx_bytes: bytes) -> bytes:
    """某些 xlsx 的 styles.xml 会让 openpyxl 崩溃（Fill() takes no arguments）。
    仅样式块受损、cell 数值无损：移除 <fills>…</fills> 并替换为默认 2-fill 骨架，
    重打包后 openpyxl 即可正常加载。"""
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zin:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    text = data.decode("utf-8", "ignore")
                    text = re.sub(
                        r"<fills\b.*?</fills>",
                        '<fills count="2">'
                        '<fill><patternFill patternType="none"/></fill>'
                        '<fill><patternFill patternType="gray125"/></fill>'
                        '</fills>',
                        text, flags=re.S,
                    )
                    data = text.encode("utf-8")
                zout.writestr(item, data)
    return buf.getvalue()


def load_workbook_patched(xlsx_bytes: bytes):
    """尽力加载工作簿；若 openpyxl 因 styles.xml 崩溃则自动修复后重试。"""
    try:
        return load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception:
        return load_workbook(io.BytesIO(_repair_xlsx_bytes(xlsx_bytes)), data_only=True)


def detect_data_sheets(wb) -> list:
    """识别工作簿中真实的 COSMIC 数据 sheet：
    第3行含标准表头（功能过程 / 子过程描述 / 数据移动类型），且第5行起至少 10 行
    I 列∈{E,W,R,X}（真实数据移动标记）。排除跟踪页/模板/空/汇总页——名称含
    「示例/模板/样例/示范」等明确非项目语义的 sheet 直接跳过。"""
    HEADER_HINTS = ("功能过程", "子过程描述", "数据移动类型")
    TEMPLATE_HINTS = ("示例", "模板", "样例", "示范", "template", "example", "sample")
    MOVES = {"E", "W", "R", "X"}
    found = []
    for ws in wb.worksheets:
        title = ws.title or ""
        if any(h in title.lower() for h in TEMPLATE_HINTS):
            continue
        row3 = " ".join(str(ws.cell(3, c).value or "") for c in range(1, 12))
        if not all(h in row3 for h in HEADER_HINTS):
            continue
        move_rows = 0
        for r in range(5, min(ws.max_row, 2000) + 1):
            v = ws.cell(r, 9).value
            if v and str(v).strip().upper()[:1] in MOVES:
                move_rows += 1
        if move_rows >= 10:
            found.append(title)
    return found


def build_bulk_mapping(dim_db: str, wb, sheets: list) -> list:
    """把识别出的数据 sheet 映射到导入项：按 requirement_name 命中存量项目则保留身份重写，
    未命中则新建（REF 顺延，status=archived）。"""
    existing = db.query(dim_db, "SELECT id, requirement_id, requirement_name FROM projects ORDER BY id")
    ex_by_name = {(p["requirement_name"] or "").strip(): p for p in existing}
    max_ref = 0
    for p in existing:
        rid = p["requirement_id"] or ""
        if rid.startswith("REF-"):
            try:
                max_ref = max(max_ref, int(rid[4:]))
            except ValueError:
                pass
    mapping = []
    for name in sheets:
        ex = ex_by_name.get(name.strip())
        if ex:
            mapping.append({"sheet": name, "project_id": ex["id"]})
        else:
            max_ref += 1
            mapping.append({"sheet": name, "project_meta": {
                "project_code": "ref_" + name.replace(" ", "")[:12],
                "client_name": "",
                "requirement_id": f"REF-{max_ref}",
                "requirement_name": name,
            }})
    return mapping


def _merged_value(ws, row, col, prev):
    v = ws.cell(row, col).value
    if v is not None:
        return str(v).strip()
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= col <= mc.max_col:
            first = ws.cell(mc.min_row, col).value
            if first is not None:
                return str(first).strip()
    return prev


def group_tree(rows: list) -> list:
    """行序 → [模块[FP[子过程]]] 树，保持文件顺序。"""
    modules, cur_m, cur_f = [], None, None
    for r in rows:
        mkey = (r["level1"], r["level2"], r["level3"])
        fkey = (r["fp_name"], r["user"], r["event"])
        if mkey != cur_m:
            cur_m, cur_f = mkey, None
            modules.append({"level1": r["level1"], "level2": r["level2"],
                            "level3": r["level3"], "fps": []})
        if fkey != cur_f:
            cur_f = fkey
            modules[-1]["fps"].append({"fp_name": r["fp_name"], "user": r["user"],
                                       "event": r["event"], "subs": []})
        modules[-1]["fps"][-1]["subs"].append(r)
    return modules


def _parse_project_meta(parsed: dict, fallback: dict | None) -> dict:
    """从 A 列标签【需求编号】需求名 解析项目信息，回退到调用方传入值。"""
    label = parsed.get("requirement_label") or ""
    m = REQ_LABEL_RE.match(label)
    rid, rname = (m.group(1), m.group(2)) if m else ("", label or "导入项目")
    fb = fallback or {}
    return {
        "project_code": fb.get("project_code") or "imported",
        "client_name": fb.get("client_name") or "",
        "requirement_id": fb.get("requirement_id") or rid,
        "requirement_name": fb.get("requirement_name") or rname,
    }


def import_xlsx(dim_db: str, xlsx_bytes: bytes, mode: str = "incremental",
                project_id: int | None = None, project_meta: dict | None = None) -> dict:
    parsed = parse_xlsx_rows(xlsx_bytes)
    rows = parsed["rows"]
    if not rows:
        raise ValueError("xlsx 无有效数据行（从第5行起读A-K列）")
    modules = group_tree(rows)
    report = {"mode": mode, "modules": len(modules),
              "fps": sum(len(m["fps"]) for m in modules),
              "subs": sum(len(f["subs"]) for m in modules for f in m["fps"]),
              "created": {"projects": 0, "modules": 0, "fps": 0, "subs": 0},
              "updated": {"modules": 0, "fps": 0}}

    with db.tx(dim_db) as cur:
        if mode == "overwrite":
            from .json_io import dump_dimension_backup
            report["backup"] = dump_dimension_backup(dim_db, tag="pre_overwrite")
            if project_id:
                cur.execute("SELECT id FROM projects WHERE id=%s", (project_id,))
                if not cur.fetchone():
                    raise ValueError(f"project_id={project_id} 不存在")
                _wipe(cur, project_id=project_id)
                target_pid = project_id
            else:
                _wipe(cur)
                meta = _parse_project_meta(parsed, project_meta)
                cur.execute("INSERT INTO projects (project_code, client_name, requirement_id, requirement_name, status) VALUES (%s,%s,%s,%s,%s)",
                            (meta["project_code"], meta["client_name"],
                             meta["requirement_id"], meta["requirement_name"], "draft"))
                target_pid = cur.lastrowid
                report["created"]["projects"] += 1
        else:
            if not project_id:
                raise ValueError("增量导入必须指定 project_id（写入哪个项目）")
            cur.execute("SELECT id FROM projects WHERE id=%s", (project_id,))
            if not cur.fetchone():
                raise ValueError(f"project_id={project_id} 不存在")
            target_pid = project_id

        _insert_tree(cur, target_pid, modules, report)
        report["project_id"] = target_pid
    return report


def _insert_tree(cur, pid: int, modules: list, report: dict) -> None:
    """把 [模块[FP[子过程]]] 树写入指定项目（模块级 upsert，FP/子过程整体重写）。"""
    for mod_order, mod in enumerate(modules, 1):
        cur.execute("SELECT id FROM modules WHERE project_id=%s AND level1=%s AND level2=%s AND level3=%s",
                    (pid, mod["level1"], mod["level2"], mod["level3"]))
        hit = cur.fetchone()
        if hit:
            mid = hit["id"]
            cur.execute("UPDATE modules SET sort_order=%s WHERE id=%s", (mod_order, mid))
            report["updated"]["modules"] += 1
        else:
            cur.execute("INSERT INTO modules (project_id, level1, level2, level3, sort_order) VALUES (%s,%s,%s,%s,%s)",
                        (pid, mod["level1"], mod["level2"], mod["level3"], mod_order))
            mid = cur.lastrowid
            report["created"]["modules"] += 1

        for fp_order, fp in enumerate(mod["fps"], 1):
            cur.execute("SELECT id FROM fps WHERE module_id=%s AND fp_name=%s", (mid, fp["fp_name"]))
            hit = cur.fetchone()
            if hit:
                fid = hit["id"]
                cur.execute("UPDATE fps SET functional_user=%s, trigger_event=%s, sort_order=%s WHERE id=%s",
                            (fp["user"], fp["event"], fp_order, fid))
                cur.execute("DELETE FROM sub_processes WHERE fp_id=%s", (fid,))
                cur.execute("DELETE FROM screenshots WHERE fp_id=%s", (fid,))
                report["updated"]["fps"] += 1
            else:
                cur.execute("INSERT INTO fps (module_id, sort_order, functional_user, trigger_event, fp_name) VALUES (%s,%s,%s,%s,%s)",
                            (mid, fp_order, fp["user"], fp["event"], fp["fp_name"]))
                fid = cur.lastrowid
                report["created"]["fps"] += 1
            for idx, sub in enumerate(fp["subs"], 1):
                cur.execute("INSERT INTO sub_processes (fp_id, sort_order, description, data_move_type, data_group_name, data_attributes) VALUES (%s,%s,%s,%s,%s,%s)",
                            (fid, idx, sub["description"], sub["move_type"],
                             sub["group_name"], sub["attributes"]))
                report["created"]["subs"] += 1


def bulk_import_workbook(dim_db: str, path: str, mapping: list,
                         backup_tag: str = "pre_bulk_overwrite") -> dict:
    """多 sheet 覆盖导入（从文件路径加载）。每个 mapping 项 = 一个项目，整维度仅备份一次。

    mapping 项格式（二选一）：
      {"sheet": <sheet名>, "project_id": <已存在项目id>}
          -> 保留项目身份（project_code/requirement_id/name 不变），仅清空并重建子数据
      {"sheet": <sheet名>, "project_meta": {project_code, client_name, requirement_id, requirement_name}}
          -> 新建项目（status=archived）
    """
    wb = load_workbook(path, data_only=True)
    try:
        return _bulk_import_workbook_wb(dim_db, wb, mapping, backup_tag)
    finally:
        wb.close()


def bulk_import_workbook_bytes(dim_db: str, xlsx_bytes: bytes,
                               backup_tag: str = "pre_bulk_workbook") -> dict:
    """多 sheet 批量覆盖导入（从上传字节加载）：自动识别数据 sheet 并映射到归档项目。"""
    wb = load_workbook_patched(xlsx_bytes)
    try:
        sheets = detect_data_sheets(wb)
        if not sheets:
            raise ValueError(
                "未识别到任何 COSMIC 数据 sheet（需第3行含 功能过程/子过程描述/数据移动类型 "
                "表头，且第5行起含 E/W/R/X 数据移动标记）")
        mapping = build_bulk_mapping(dim_db, wb, sheets)
        return _bulk_import_workbook_wb(dim_db, wb, mapping, backup_tag)
    finally:
        wb.close()


def _bulk_import_workbook_wb(dim_db: str, wb, mapping: list,
                             backup_tag: str = "pre_bulk_overwrite") -> dict:
    """多 sheet 覆盖导入核心：整维度仅备份一次，逐 mapping 项清子数据重建（保留/新建项目身份）。"""
    from .json_io import dump_dimension_backup
    report = {"mode": "bulk_overwrite", "sheets": len(mapping),
              "created": {"projects": 0, "modules": 0, "fps": 0, "subs": 0},
              "updated": {"modules": 0, "fps": 0}, "skipped": [], "errors": [], "details": []}
    with db.tx(dim_db) as cur:
        report["backup"] = dump_dimension_backup(dim_db, tag=backup_tag)
        for item in mapping:
            name = item["sheet"]
            if name not in wb.sheetnames:
                report["errors"].append({"sheet": name, "error": "sheet 不存在"})
                continue
            parsed = parse_worksheet(wb[name])
            rows = parsed["rows"]
            if not rows:
                report["skipped"].append({"sheet": name, "reason": "无数据行"})
                continue
            modules = group_tree(rows)
            if "project_id" in item:
                pid = item["project_id"]
                cur.execute("SELECT id FROM projects WHERE id=%s", (pid,))
                if not cur.fetchone():
                    report["errors"].append({"sheet": name, "error": f"project_id={pid} 不存在"})
                    continue
                _wipe(cur, project_id=pid)
            else:
                meta = item.get("project_meta") or _parse_project_meta(parsed, None)
                cur.execute("INSERT INTO projects (project_code, client_name, requirement_id, requirement_name, status) VALUES (%s,%s,%s,%s,%s)",
                            (meta["project_code"], meta["client_name"],
                             meta["requirement_id"], meta["requirement_name"], "archived"))
                pid = cur.lastrowid
                report["created"]["projects"] += 1
            _insert_tree(cur, pid, modules, report)
            report["details"].append({
                "sheet": name, "project_id": pid,
                "modules": len(modules),
                "fps": sum(len(m["fps"]) for m in modules),
                "subs": sum(len(f["subs"]) for m in modules for f in m["fps"]),
            })
    report["modules"] = sum(d["modules"] for d in report["details"])
    report["fps"] = sum(d["fps"] for d in report["details"])
    report["subs"] = sum(d["subs"] for d in report["details"])
    report["projects"] = len(report["details"])
    return report


def _wipe(cur, project_id: int | None = None):
    """清数据（整维度或单项目）。项目级覆盖只清数据保留项目壳，防孤儿模块。"""
    if project_id:
        cur.execute("DELETE FROM screenshots WHERE fp_id IN (SELECT f.id FROM fps f JOIN modules m ON m.id=f.module_id WHERE m.project_id=%s)", (project_id,))
        cur.execute("DELETE FROM sub_processes WHERE fp_id IN (SELECT f.id FROM fps f JOIN modules m ON m.id=f.module_id WHERE m.project_id=%s)", (project_id,))
        cur.execute("DELETE FROM fps WHERE module_id IN (SELECT id FROM modules WHERE project_id=%s)", (project_id,))
        cur.execute("DELETE FROM modules WHERE project_id=%s", (project_id,))
    else:
        for t in ("screenshots", "sub_processes", "fps", "modules", "projects"):
            cur.execute(f"DELETE FROM {t}")
