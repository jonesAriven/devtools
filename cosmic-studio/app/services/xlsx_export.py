"""标准 COSMIC xlsx 导出：逐字节对照 Hermes generate_xlsx.py 的固定格式。

格式常量写死不改（列宽/行高/字体/填充/边框/合并模式）：
  表头 4 行 + 数据从第 5 行起，行高 60
  A-C 跨整个需求合并；D 按模块合并；E/F/G 按功能过程合并；L-M 按功能过程合并；
  H/I/J/K 每行独立（J 数据组不合并）
  每行写全 A-P 列值再 merge（合并区子过程行不可留空）
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .. import config, db

COL_WIDTHS = {
    "A": 27.42, "B": 25.22, "C": 24.79, "D": 26.46,
    "E": 22.02, "F": 38.65, "G": 35.19, "H": 52.36,
    "I": 16.49, "J": 43.64, "K": 84.64, "L": 30.0,
    "M": 30.0, "N": 15.0, "O": 20.0, "P": 12.0,
}
ROW_HEIGHTS = {1: 22.05, 2: 23.85, 3: 19.7, 4: 19.7}
FONT_TITLE = Font(name="Noto Sans CJK SC", size=18, bold=True)
FONT_HEADER = Font(name="Noto Sans CJK SC", size=16, bold=True)
FONT_DATA = Font(name="Noto Sans CJK SC", size=11)
FONT_I_COL = Font(name="宋体", size=11)
FILL_GRAY = PatternFill(start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid")
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_DATA = Alignment(horizontal="general", vertical="center", wrap_text=True)
ALIGN_I_COL = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
COL = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7,
       "H": 8, "I": 9, "J": 10, "K": 11, "L": 12, "M": 13,
       "N": 14, "O": 15, "P": 16}


def _apply_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = THIN_BORDER


def load_project_tree(dim_db: str, project_id: int):
    """读出 项目→模块→FP→子过程 结构化树（与导出/JSON 共用）。"""
    proj = db.query(dim_db, "SELECT * FROM projects WHERE id=%s", (project_id,), one=True)
    if not proj:
        return None
    mods = db.query(dim_db, "SELECT * FROM modules WHERE project_id=%s ORDER BY sort_order", (project_id,))
    tree = dict(proj)
    tree["modules"] = []
    for mod in mods:
        m = dict(mod)
        m["fps"] = []
        fps = db.query(dim_db, "SELECT * FROM fps WHERE module_id=%s ORDER BY sort_order", (mod["id"],))
        for fp in fps:
            f = dict(fp)
            f["subs"] = db.query(dim_db,
                                 "SELECT * FROM sub_processes WHERE fp_id=%s ORDER BY sort_order", (fp["id"],))
            m["fps"].append(f)
        tree["modules"].append(m)
    return tree


def export_xlsx(dim_db: str, project_id: int, author: str = "") -> tuple[bytes, dict]:
    tree = load_project_tree(dim_db, project_id)
    if not tree:
        raise ValueError(f"project_id={project_id} not found")
    if not author:
        author = config.DEFAULT_INITIATOR

    module_list = tree["modules"]
    if not any(fp["subs"] for mod in module_list for fp in mod["fps"]):
        raise ValueError("项目无子过程数据，无法导出")

    wb = Workbook()
    ws = wb.active
    ws.title = "COSMIC"

    total_data_rows = sum(len(fp["subs"]) for mod in module_list for fp in mod["fps"])
    first_row, last_row = 5, 5 + total_data_rows - 1
    max_col = 16

    # 表头 1-4 行
    ws.merge_cells("A1:P1")
    c = ws.cell(1, 1, "通用软件评估模型")
    c.font, c.fill, c.alignment = FONT_TITLE, FILL_GRAY, ALIGN_HEADER
    for rng, text in (("A2:E2", "度量策略阶段"), ("F2:K2", "映射阶段"), ("L2:P2", "度量阶段")):
        ws.merge_cells(rng)
        cell = ws.cell(2, COL[rng[0]], text)
        cell.font, cell.fill, cell.alignment = FONT_HEADER, FILL_GRAY, ALIGN_HEADER
    headers_3 = {"A": "客户需求", "B": "功能用户需求", "E": "功能用户", "F": "触发事件",
                 "G": "功能过程", "H": "子过程描述", "I": "数据移动类型", "J": "数据组",
                 "K": "数据属性", "L": "功能过程截图\n（可以放多张）",
                 "N": "cosmic编写人", "O": "评审意见", "P": "是否修改"}
    for col, text in headers_3.items():
        c = ws.cell(3, COL[col], text)
        c.font, c.fill, c.alignment = FONT_HEADER, FILL_GRAY, ALIGN_HEADER
    ws.merge_cells("B3:D3")
    for col in ["A", "E", "F", "G", "H", "I", "J", "K", "N", "O", "P"]:
        ws.merge_cells(f"{col}3:{col}4")
    ws.merge_cells("L3:M4")
    for col, text in {"B": "一级模块", "C": "二级模块", "D": "三级模块"}.items():
        c = ws.cell(4, COL[col], text)
        c.font, c.fill, c.alignment = FONT_HEADER, FILL_GRAY, ALIGN_HEADER
    _apply_border(ws, 1, 4, 1, max_col)
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    for row_num, height in ROW_HEIGHTS.items():
        ws.row_dimensions[row_num].height = height

    # N 列编写人：写在 FP 首行（后随 E/F/G 合并逻辑独立，N 列每 FP 一值）
    row = first_row
    for mod in module_list:
        d_start = row
        for fp in mod["fps"]:
            e_start = row
            ws.cell(row, COL["N"], author)
            for sub in fp["subs"]:
                a_val = f"【{tree['requirement_id']}】{tree['requirement_name']}"
                for col, val in (("A", a_val), ("B", mod["level1"]), ("C", mod["level2"]),
                                 ("D", mod["level3"]), ("E", fp["functional_user"]),
                                 ("F", fp["trigger_event"]), ("G", fp["fp_name"]),
                                 ("H", sub["description"]), ("I", sub["data_move_type"]),
                                 ("J", sub["data_group_name"]), ("K", sub["data_attributes"])):
                    ws.cell(row, COL[col], val)
                for ci in range(1, max_col + 1):
                    cell = ws.cell(row, ci)
                    cell.font, cell.alignment, cell.border = FONT_DATA, ALIGN_DATA, THIN_BORDER
                ws.cell(row, COL["I"]).font = FONT_I_COL
                ws.cell(row, COL["I"]).alignment = ALIGN_I_COL
                ws.row_dimensions[row].height = 60.0
                row += 1
            fp_end = row - 1
            if fp_end > e_start:
                for col in ("E", "F", "G"):
                    ws.merge_cells(f"{col}{e_start}:{col}{fp_end}")
                ws.merge_cells(f"L{e_start}:M{fp_end}")
                ws.merge_cells(f"N{e_start}:N{fp_end}")
        d_end = row - 1
        if d_end > d_start:
            ws.merge_cells(f"D{d_start}:D{d_end}")
    if last_row > first_row:
        for col in ("A", "B", "C"):
            ws.merge_cells(f"{col}{first_row}:{col}{last_row}")

    meta = {"rows": total_data_rows, "modules": len(module_list),
            "fps": sum(len(m["fps"]) for m in module_list)}
    _embed_screenshots(dim_db, ws, tree, first_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), meta


def build_import_template() -> bytes:
    """导入模板：COSMIC 表头 + 灰色示例行 + 「填写说明」sheet。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font as _F

    wb = Workbook()
    ws = wb.active
    ws.title = "COSMIC"
    max_col = 16

    ws.merge_cells("A1:P1")
    c = ws.cell(1, 1, "通用软件评估模型（导入模板）")
    c.font, c.fill, c.alignment = FONT_TITLE, FILL_GRAY, ALIGN_HEADER
    for rng, text in (("A2:E2", "度量策略阶段"), ("F2:K2", "映射阶段"), ("L2:P2", "度量阶段")):
        ws.merge_cells(rng)
        cell = ws.cell(2, COL[rng[0]], text)
        cell.font, cell.fill, cell.alignment = FONT_HEADER, FILL_GRAY, ALIGN_HEADER
    headers_3 = {"A": "客户需求", "B": "功能用户需求", "E": "功能用户", "F": "触发事件",
                 "G": "功能过程", "H": "子过程描述", "I": "数据移动类型", "J": "数据组",
                 "K": "数据属性", "L": "功能过程截图\n（可以放多张）",
                 "N": "cosmic编写人", "O": "评审意见", "P": "是否修改"}
    for col, text in headers_3.items():
        c = ws.cell(3, COL[col], text)
        c.font, c.fill, c.alignment = FONT_HEADER, FILL_GRAY, ALIGN_HEADER
    ws.merge_cells("B3:D3")
    for col in ["A", "E", "F", "G", "H", "I", "J", "K", "N", "O", "P"]:
        ws.merge_cells(f"{col}3:{col}4")
    ws.merge_cells("L3:M4")
    for col, text in {"B": "一级模块", "C": "二级模块", "D": "三级模块"}.items():
        c = ws.cell(4, COL[col], text)
        c.font, c.fill, c.alignment = FONT_HEADER, FILL_GRAY, ALIGN_HEADER
    _apply_border(ws, 1, 4, 1, max_col)
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    for row_num, height in ROW_HEIGHTS.items():
        ws.row_dimensions[row_num].height = height

    # 示例数据行（第5行，灰色斜体，提示导入前删除或直接覆盖）
    demo = ["【202607270156】互联网卡片-已订购业务卡片支持展示资费标准",
            "互联网卡片", "卡片管理", "资费标准配置管理",
            "发起者：一线坐席\n接收者：多媒体卡片平台",
            "一线坐席新增资费标准显隐开关时触发", "新增资费标准显隐开关",
            "接收一线坐席发起新增资费标准显隐开关请求", "E",
            "资费标准显隐开关新增请求数据",
            "开关编号、卡片类型编码、目标省份编码、显隐状态"]
    f_demo = Font(name="Noto Sans CJK SC", size=11, italic=True, color="FF808080")
    for ci, val in enumerate(demo, 1):
        cell = ws.cell(5, ci, val)
        cell.font, cell.alignment, cell.border = f_demo, ALIGN_DATA, THIN_BORDER
        if ci == 9:
            cell.font = Font(name="宋体", size=11, italic=True, color="FF808080")
            cell.alignment = ALIGN_I_COL
    ws.row_dimensions[5].height = 60.0

    # 填写说明 sheet
    guide = wb.create_sheet("填写说明")
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 100
    rows = [
        ("列", "填写要求"),
        ("数据起始行", "第 5 行起；L-P 列为展示列（截图/编写人/评审），导入时不读取"),
        ("B/C/D 列", "一级/二级/三级模块名。三级模块名禁含禁词（记录、日志、导入、缓存、明细、列表、详情、效果）"),
        ("E 列", "功能用户，固定两行格式：第一行「发起者：XX」，第二行「接收者：XX」"),
        ("F 列", "触发事件，格式「{发起者}{功能过程名}时触发」，不含接收者、不换行"),
        ("G 列", "功能过程名，动词开头（新增/修改/删除/查询/预览 + 业务对象）"),
        ("H 列", "子过程描述，E 类以「接收」开头、X 类以「返回」开头，不含逗号断句"),
        ("I 列", "数据移动类型：E(输入)/W(写入)/R(读取)/X(输出)。新增/修改/删除类 FP 用 EW，查询/预览类用 ERX"),
        ("J 列", "数据组名，每子过程独立不合并。E 类以「请求数据」结尾，W 类以「数据」结尾，R 类以「查询数据」结尾，X 类以「查询结果」结尾"),
        ("K 列", "数据属性，用「、」分隔（禁用逗号），每行至少 3 个字段（建议≥4）；字段须为真实数据库列名，禁行为词/统计值/PII（客户姓名、证件号等）"),
        ("增量导入", "按业务主键 upsert：模块=(一级+二级+三级)，功能过程=(模块+G列名)。命中的 FP 更新并整体重写其子过程，未命中新建；不触碰库内其他数据"),
        ("覆盖导入", "清空目标（所选项目或整库）后按文件重灌；导入前系统自动备份现有数据到 /data/backups/；需 admin 权限"),
        ("导入后建议", "到项目详情页点「推导检查」和「质量门禁」，errors=0 再交付；格式类问题可用「一键修复推导列」自动修正"),
    ]
    for ri, (a, b) in enumerate(rows, 1):
        ca = guide.cell(ri, 1, a)
        cb = guide.cell(ri, 2, b)
        ca.font = _F(name="Noto Sans CJK SC", size=11, bold=(ri == 1))
        cb.font = _F(name="Noto Sans CJK SC", size=11, bold=(ri == 1))
        cb.alignment = ALIGN_DATA
        guide.row_dimensions[ri].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _embed_screenshots(dim_db, ws, tree, first_row):
    """FP 有截图记录时嵌入 L 列（每 FP 取第一张）。"""
    total = 0
    row = first_row
    for mod in tree["modules"]:
        for fp in mod["fps"]:
            img = db.query(dim_db, "SELECT image_data, image_width, image_height FROM screenshots "
                                   "WHERE fp_id=%s ORDER BY sort_order LIMIT 1", (fp["id"],), one=True)
            if img:
                import io as _io
                xl = XLImage(_io.BytesIO(img["image_data"]))
                xl.width = min(img["image_width"], 500)
                xl.height = min(img["image_height"], 300)
                ws.add_image(xl, f"L{row}")
                total += 1
                target = max(60, xl.height + 10)
                for r in range(row, row + len(fp["subs"])):
                    ws.row_dimensions[r].height = target
            row += len(fp["subs"])
    return total
